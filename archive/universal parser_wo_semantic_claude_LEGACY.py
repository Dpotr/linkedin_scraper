#!/usr/bin/env python3
import logging
import time
import io
import threading
import tkinter as tk
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import pandas as pd
import requests
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from langdetect import detect, LangDetectException
import undetected_chromedriver as uc
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
import re
import os
from datetime import datetime

# ================================
# Настройка логирования
# ================================
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ================================
# Значения по умолчанию
# ================================
DEFAULT_TELEGRAM_BOT_TOKEN = "7690052120:AAHewK4ztdFw7y-iCApCmRMkCwz9inLLkfI"
DEFAULT_TELEGRAM_CHAT_ID = "1908191"
DEFAULT_CHROMEDRIVER_PATH = r"C:\selenium\chromedriver.exe"
DEFAULT_CHROME_PROFILE_PATH = r"C:\Users\potre\SeleniumProfileNew"
DEFAULT_CHROME_BINARY_LOCATION = r"C:\Program Files\Google\Chrome Beta\Application\chrome.exe"
DEFAULT_OUTPUT_FILE_PATH = r"C:\Users\potre\OneDrive\LinkedIn_Automation\companies_usa_remote.xlsx"
DEFAULT_SEARCH_COUNTRY = "United States"
DEFAULT_KEYWORD = "remote job"

# ================================
# Ключевые слова для поиска - используются в URL LinkedIn
# ================================
SEARCH_KEYWORDS = [
    "remote job",
    "supply chain planner remote",
    "anaplan remote",
    "sap apo remote",
    "demand planning remote",
    "supply planning remote"
]

# ================================
# Ключевые слова для фильтрации описаний работы
# ================================
NO_RELOCATION_REQUIREMENTS = [
    "no relocation support", "we do not provide relocation", "no visa sponsorship",
    "relocation not provided", "visa sponsorship unavailable", "relocation assistance not offered",
    "visa sponsorship not included", "no relocation assistance", "we are unable to sponsor visas",
    "visa sponsorship not possible", "local applicants only", "we cannot provide relocation"
]
REMOTE_REQUIREMENTS = [
    "remote", "work from home", "fully remote", "telecommute", "telecommuting", "remote work", 
    "remote position", "virtual", "work anywhere", "flexible location"
]
KEYWORDS_VISA = [
    "we provide relocation", "relocation assistance", "relocation support", "we sponsor visas",
    "visa sponsorship available", "work visa sponsorship", "relocation package", "we support relocation",
    "visa sponsorship provided", "relocation offered", "visa and relocation assistance",
    "international relocation", "relocation and visa support"
]
KEYWORDS_ANAPLAN = [
    "anaplan", "anaplan model builder", "anaplan planning", "anaplan developer", "anaplan consultant",
    "anaplan architect", "anaplan implementation", "anaplan solution", "anaplan support", "anaplan admin",
    "connected planning", "anaplan integration"
]
KEYWORDS_SAP = [
    "sap snp", "sap apo", "sap pp/ds", "sap scm", "sap supply planning", "sap advanced planning and optimization",
    "sap production planning", "sap demand planning", "sap distribution planning", "sap key user",
    "sap supply chain management"
]
KEYWORDS_PLANNING = [
    "supply planning", "supply chain planning", "distribution requirements planning", "inventory planning", 
    "production planning", "demand planning", "material requirements planning", "capacity planning", 
    "supply chain planning", "logistics planning", "operations planning", "master production scheduling", 
    "forecasting", "s&op", "sales and operations planning", "warehouse planning", "network planning", 
    "procurement planning"
]
COMMON_SKILLS = [
    "python", "sql", "excel", "power bi", "tableau", "data analysis", "statistics", 
    "machine learning", "optimization", "modeling", "simulation", "forecasting"
]
RED_FLAGS = [
    "unpaid", "commission only", "unlimited earning potential", "be your own boss",
    "work hard play hard", "ninja", "rockstar", "guru", "fast-paced environment"
]

# Все ключевые слова для проверки
ALL_KEYWORDS = (
    KEYWORDS_VISA
    + KEYWORDS_ANAPLAN
    + KEYWORDS_SAP
    + KEYWORDS_PLANNING
    + NO_RELOCATION_REQUIREMENTS
    + REMOTE_REQUIREMENTS
    + COMMON_SKILLS
)

# Глобальные переменные
results = []
total_vacancies_checked = 0

# ================================
# Функция отправки сообщений в Telegram
# ================================
def send_telegram_message(bot_token, chat_id, message, images=None):
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    data = {"chat_id": chat_id, "text": message, "parse_mode": "HTML"}
    try:
        resp = requests.post(url, data=data)
        if resp.status_code == 200:
            logging.info("Сообщение отправлено в Telegram.")
        else:
            logging.error(f"Ошибка при отправке сообщения: {resp.text}")
    except Exception as e:
        logging.error(f"Ошибка подключения к Telegram: {e}")
    
    if images:
        for img in images:
            url_photo = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
            files = {"photo": img}
            data = {"chat_id": chat_id}
            try:
                resp_img = requests.post(url_photo, files=files, data=data)
                if resp_img.status_code == 200:
                    logging.info("Изображение отправлено в Telegram.")
                else:
                    logging.error(f"Ошибка при отправке изображения: {resp_img.text}")
            except Exception as e:
                logging.error(f"Ошибка подключения к Telegram при отправке изображения: {e}")

# ================================
# Работа с Excel
# ================================
def get_excel_summary(file_path, running_time_minutes):
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        summary = []
        summary.append(f"Running time, min: {running_time_minutes:.2f}")
        summary.append(f"Checked companies: {sheet.cell(row=1, column=19).value} {sheet.cell(row=1, column=20).value}")
        return "\n".join(summary)
    except Exception as e:
        logging.error(f"Ошибка при извлечении данных из Excel: {e}")
        return ""

def save_results_to_file_with_calculations(results, output_file, elapsed_time):
    try:
        df = pd.DataFrame(results)
        
        # Добавляем дополнительные столбцы для аналитики
        if not df.empty:
            # Подсчет совпадений по категориям
            true_count = sum(
                any(value is True for key, value in row.items() if key not in ["Elapsed Time (s)", "Job URL", "Red Flags", "Skills"])
                for row in results
            )
            ratio = true_count / len(df) if len(df) > 0 else 0
            processed_companies = len(df)
            avg_time_per_company = elapsed_time / processed_companies if processed_companies > 0 else 0
            
            # Добавляем столбец с датой/временем
            df["Processed At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Создаем или добавляем к существующему файлу
        try:
            existing_df = pd.read_excel(output_file)
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df.to_excel(output_file, index=False)
        except FileNotFoundError:
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Results")
                wb = writer.book
                sheet = writer.sheets["Results"]
                sheet["S1"] = "Checked companies:"
                sheet["T1"] = processed_companies

        logging.info(f"Результаты сохранены в файл: {output_file}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении результатов в Excel: {e}")

# ================================
# Построение аналитики
# ================================
def create_p_chart(results):
    if len(results) < 2:
        return None
    df = pd.DataFrame(results)
    try:
        df["Elapsed Time (s)"] = df["Elapsed Time (s)"].astype(float)
        df["Time Difference (s)"] = df["Elapsed Time (s)"].diff()
        time_diffs = df["Time Difference (s)"][1:].tail(50)
        avg_time_diff = time_diffs.mean()
        plt.figure(figsize=(10, 5))
        plt.plot(time_diffs.index, time_diffs, marker='o', linestyle='-', color='b', label='Time Difference (s)')
        plt.axhline(y=avg_time_diff, color='r', linestyle='--', label=f'Avg Diff ({avg_time_diff:.2f}s)')
        plt.fill_between(time_diffs.index, avg_time_diff, time_diffs,
                         where=(time_diffs > avg_time_diff), color='red', alpha=0.1)
        plt.fill_between(time_diffs.index, avg_time_diff, time_diffs,
                         where=(time_diffs < avg_time_diff), color='green', alpha=0.1)
        plt.xlabel('Listing Index')
        plt.ylabel('Time Diff (s)')
        plt.title('P-Chart: Time Difference (Last 50)')
        plt.legend()
        plt.grid(True)
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close()
        return buf
    except Exception as e:
        logging.error(f"Ошибка при построении p-chart: {e}")
        return None

def create_bar_chart(results):
    if not results:
        return None
    df = pd.DataFrame(results)
    criteria = ["Visa Sponsorship or Relocation", "Anaplan", "SAP APO", "Planning", "No Relocation Support", "Remote", "Skills"]
    try:
        counts = {crit: df[crit].sum() if crit in df.columns else 0 for crit in criteria}
        plt.figure(figsize=(10, 6))
        colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6', '#1abc9c', '#34495e']
        bars = plt.bar(counts.keys(), counts.values(), color=colors)
        
        # Добавление значений над столбцами
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                    f'{int(height)}', ha='center', va='bottom')
        
        plt.xlabel('Критерии', fontsize=12)
        plt.ylabel('Количество', fontsize=12)
        plt.title('Распределение критериев в обработанных вакансиях', fontsize=14)
        num_labels = len(counts)
        fontsize = max(6, 12 - num_labels // 2)
        plt.xticks(rotation=45, fontsize=fontsize)
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close()
        return buf
    except Exception as e:
        logging.error(f"Ошибка при построении бар-чарта: {e}")
        return None

def create_skills_chart(results):
    """Создает график с упоминаемыми навыками"""
    if not results:
        return None
    
    # Подсчет упоминаний навыков
    skill_counts = {}
    for result in results:
        if "Skills" in result and result["Skills"]:
            for skill in result["Skills"]:
                skill_counts[skill] = skill_counts.get(skill, 0) + 1
    
    if not skill_counts:
        return None
    
    # Сортировка по количеству упоминаний
    sorted_skills = sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)
    top_skills = sorted_skills[:10]  # Топ-10 навыков
    
    try:
        plt.figure(figsize=(10, 6))
        skills = [item[0] for item in top_skills]
        counts = [item[1] for item in top_skills]
        
        # Используем горизонтальные полосы для лучшей читаемости
        bars = plt.barh(skills, counts, color='#2ecc71')
        
        # Добавление значений в конце полос
        for i, bar in enumerate(bars):
            width = bar.get_width()
            plt.text(width + 0.3, bar.get_y() + bar.get_height()/2, 
                    f'{int(width)}', ha='left', va='center')
        
        plt.xlabel('Количество упоминаний', fontsize=12)
        plt.title('Топ-10 запрашиваемых навыков', fontsize=14)
        plt.tight_layout()
        plt.grid(axis='x', linestyle='--', alpha=0.7)
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close()
        return buf
    except Exception as e:
        logging.error(f"Ошибка при построении графика навыков: {e}")
        return None

# ================================
# Функция прокрутки
# ================================
def scroll_until_loaded(driver, pause_time=1, max_consecutive=3):
    global total_vacancies_checked
    consecutive = 0
    prev_count = 0
    body = driver.find_element(By.TAG_NAME, "body")
    # Принудительный PAGE_DOWN на первой странице
    body.send_keys(Keys.PAGE_DOWN)
    time.sleep(pause_time)
    while consecutive < max_consecutive:
        body.send_keys(Keys.PAGE_DOWN)
        time.sleep(pause_time)
        current_count = len(driver.find_elements(By.CSS_SELECTOR, ".job-card-container--clickable"))
        logging.debug(f"PAGE_DOWN -> вакансий сейчас: {current_count}")
        if current_count <= prev_count:
            consecutive += 1
        else:
            consecutive = 0
            prev_count = current_count

# ================================
# Дополнительные функции анализа вакансий
# ================================
def extract_salary_info(desc_text):
    """Извлекает информацию о зарплате из описания вакансии"""
    patterns = [
        r'\$(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*-\s*\$(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',  # $X - $Y
        r'(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*K\s*-\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*K',  # XK - YK
        r'salary range.*?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)(?:k|K)?\s*-\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)(?:k|K)?',  # salary range X-Y
        r'range.*?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*-\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',  # range X-Y
        r'pay range.*?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*-\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)',  # pay range X-Y
        r'compensation.*?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*-\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)'  # compensation X-Y
    ]
    
    desc_text_lower = desc_text.lower()
    
    for pattern in patterns:
        matches = re.findall(pattern, desc_text_lower)
        if matches:
            # Take the first match
            min_salary, max_salary = matches[0]
            
            # Remove commas
            min_salary = min_salary.replace(',', '')
            max_salary = max_salary.replace(',', '')
            
            # Convert to float
            try:
                min_value = float(min_salary)
                max_value = float(max_salary)
                
                # If values are likely in thousands (less than 1000)
                if min_value < 1000 and 'k' in desc_text_lower:
                    min_value *= 1000
                    max_value *= 1000
                
                return f"${min_value:,.0f} - ${max_value:,.0f}"
            except ValueError:
                continue
    
    return None

def extract_experience_level(desc_text):
    """Определяет предполагаемый уровень опыта из описания вакансии"""
    desc_text_lower = desc_text.lower()
    
    entry_keywords = ["entry level", "junior", "0-1 years", "0-2 years", "no experience", "recent graduate", "beginner"]
    mid_keywords = ["mid level", "intermediate", "2-5 years", "3-5 years", "some experience"]
    senior_keywords = ["senior", "5+ years", "7+ years", "10+ years", "experienced", "lead", "principal"]
    
    # Подсчет встречаемости ключевых слов каждого уровня
    entry_count = sum(1 for kw in entry_keywords if kw in desc_text_lower)
    mid_count = sum(1 for kw in mid_keywords if kw in desc_text_lower)
    senior_count = sum(1 for kw in senior_keywords if kw in desc_text_lower)
    
    # Определение по большинству упоминаний
    if senior_count > mid_count and senior_count > entry_count:
        return "Senior"
    elif mid_count > entry_count:
        return "Mid-level"
    else:
        return "Entry-level"

def extract_mentioned_skills(desc_text):
    """Извлекает упомянутые навыки из описания вакансии"""
    desc_text_lower = desc_text.lower()
    mentioned_skills = [skill for skill in COMMON_SKILLS if skill in desc_text_lower]
    return mentioned_skills

def check_red_flags(desc_text):
    """Проверяет наличие потенциальных "красных флагов" в описании вакансии"""
    desc_text_lower = desc_text.lower()
    flags = [flag for flag in RED_FLAGS if flag in desc_text_lower]
    return flags

# ================================
# Обработка вакансий на странице
# ================================
def parse_current_page(driver, wait, start_time, config):
    global results, total_vacancies_checked
    try:
        scroll_until_loaded(driver, pause_time=1, max_consecutive=3)
        job_listings = driver.find_elements(By.CSS_SELECTOR, ".job-card-container--clickable")
        logging.info(f"Найдено вакансий на странице: {len(job_listings)}")
        total_vacancies_checked += len(job_listings)
        matching_jobs = []

        for i, job in enumerate(job_listings, start=1):
            try:
                action = ActionChains(driver)
                # Получаем данные вакансии
                try:
                    job_company_name = job.find_element(By.CSS_SELECTOR, ".artdeco-entity-lockup__subtitle span").text.strip()
                except Exception as e:
                    try:
                        job_company_name = job.find_element(By.CSS_SELECTOR, ".job-card-container__company-name").text.strip()
                    except:
                        logging.debug(f"Не удалось получить имя компании: {e}")
                        job_company_name = "Unknown Company"
                
                try:
                    job_title = job.find_element(By.CSS_SELECTOR, ".artdeco-entity-lockup__title span").text.strip()
                except Exception as e:
                    try:
                        job_title = job.find_element(By.CSS_SELECTOR, ".job-card-list__title").text.strip()
                    except Exception as e2:
                        logging.debug(f"Не удалось получить заголовок вакансии: {e2}")
                        job_title = "Title not found"

                # Извлекаем местоположение, если доступно
                location = ""
                try:
                    location = job.find_element(By.CSS_SELECTOR, ".job-card-container__metadata-item").text.strip()
                except:
                    pass

                logging.info(f"Обработка вакансии №{i}: '{job_title}' / {job_company_name}")
                action.move_to_element(job).click().perform()

                # Ждем загрузки описания вакансии
                try:
                    desc_element = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "jobs-box__html-content"))
                    )
                except Exception:
                    try:
                        desc_element = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "jobs-description__content"))
                        )
                    except:
                        logging.info("Описание не удалось найти, пропускаем.")
                        continue

                # Пытаемся получить текст описания
                desc_text = ""
                for _ in range(5):  # Пробуем несколько раз
                    time.sleep(0.8)
                    tmp = desc_element.get_attribute("innerText").strip()
                    if len(tmp) > 50:
                        desc_text = tmp.lower()
                        break
                
                if not desc_text:
                    logging.info("Описание не успело прогрузиться, пропускаем.")
                    continue

                # Определяем язык описания
                try:
                    detected_language = detect(desc_text)
                except LangDetectException:
                    detected_language = "unknown"

                # Получаем URL вакансии
                job_url = driver.current_url

                # Извлекаем дополнительную информацию
                salary_info = extract_salary_info(desc_text)
                experience_level = extract_experience_level(desc_text)
                mentioned_skills = extract_mentioned_skills(desc_text)
                red_flags = check_red_flags(desc_text)

                # Определяем флаги соответствия
                remote_found = any(x in desc_text for x in REMOTE_REQUIREMENTS)
                visa_or_relocation = any(x in desc_text for x in KEYWORDS_VISA)
                anaplan_found = any(x in desc_text for x in KEYWORDS_ANAPLAN)
                sap_apo_found = any(x in desc_text for x in KEYWORDS_SAP)
                planning_found = any(x in desc_text for x in KEYWORDS_PLANNING)
                no_relocation = any(x in desc_text for x in NO_RELOCATION_REQUIREMENTS)
                already_applied = any(x in desc_text for x in ["applied", "see application", "you have already applied", "already submitted", "previously applied"])
                has_skills = bool(mentioned_skills)

                # Улучшенная логика совпадения
                is_match = not already_applied and (
                    (remote_found or visa_or_relocation) and  # Должно быть указание на удаленку или релокацию
                    (anaplan_found or sap_apo_found or planning_found)  # И одно из ключевых слов по планированию
                )

                # Если есть совпадение, сохраняем вакансию
                if is_match:
                    matched_keywords = [kw for kw in ALL_KEYWORDS if kw in desc_text]
                    
                    # Создаем результат с дополнительными данными
                    current_result = {
                        "Company": job_company_name,
                        "Vacancy Title": job_title,
                        "Location": location,
                        "Visa Sponsorship or Relocation": visa_or_relocation,
                        "Anaplan": anaplan_found,
                        "SAP APO": sap_apo_found,
                        "Planning": planning_found,
                        "No Relocation Support": no_relocation,
                        "Remote": remote_found,
                        "Skills": mentioned_skills,
                        "Has Skills": has_skills,
                        "Red Flags": red_flags,
                        "Salary Range": salary_info,
                        "Experience Level": experience_level,
                        "Already Applied": already_applied,
                        "Language": detected_language,
                        "Job URL": job_url,
                        "Elapsed Time (s)": round(time.perf_counter() - start_time, 2)
                    }
                    matching_jobs.append(current_result)
                    
                    # Подготавливаем сообщение для Telegram с дополнительной информацией
                    running_minutes = (time.perf_counter() - start_time) / 60
                    summary = get_excel_summary(config["output_file_path"], running_minutes)
                    
                    # Форматирование расширенного сообщения
                    message_text = (
                        f"🔔 <b>{job_title}</b> в компании <b>{job_company_name}</b>\n\n"
                        f"Местоположение: {location or 'Не указано'}\n"
                    )
                    
                    # Добавляем информацию о зарплате, если есть
                    if salary_info:
                        message_text += f"Зарплата: <b>{salary_info}</b>\n"
                    
                    # Добавляем информацию об опыте
                    message_text += f"Уровень опыта: <b>{experience_level}</b>\n\n"
                    
                    # Добавляем информацию о навыках
                    if mentioned_skills:
                        message_text += f"Требуемые навыки: <b>{', '.join(mentioned_skills)}</b>\n\n"
                    
                    # Добавляем совпадения по категориям
                    categories = []
                    if remote_found: categories.append("Удаленная работа")
                    if visa_or_relocation: categories.append("Релокация/Виза")
                    if anaplan_found: categories.append("Anaplan")
                    if sap_apo_found: categories.append("SAP APO")
                    if planning_found: categories.append("Планирование")
                    
                    message_text += f"Категории: <b>{', '.join(categories)}</b>\n\n"
                    
                    # Добавляем "красные флаги", если есть
                    if red_flags:
                        message_text += f"⚠️ Обратите внимание: {', '.join(red_flags)}\n\n"
                    
                    # Добавляем информацию о языке
                    message_text += f"Язык описания: {detected_language.upper()}\n\n"
                    
                    # Добавляем ссылку и статистику
                    message_text += (
                        f"<a href='{job_url}'>Открыть вакансию на LinkedIn</a>\n\n"
                        f"Всего проверено вакансий: {total_vacancies_checked}\n"
                        + summary
                    )
                    
                    # Создаем графики для Telegram
                    p_chart = create_p_chart(results)
                    bar_chart = create_bar_chart(results)
                    skills_chart = create_skills_chart(results)
                    
                    images = []
                    if p_chart:
                        images.append(p_chart)
                    if bar_chart:
                        images.append(bar_chart)
                    if skills_chart:
                        images.append(skills_chart)
                    
                    send_telegram_message(config["telegram_bot_token"], config["telegram_chat_id"], message_text, images=images)
            except Exception as e:
                logging.error(f"Ошибка при обработке вакансии №{i}: {e}")
                continue

        results.extend(matching_jobs)
    except Exception as e:
        logging.error(f"Ошибка при разборе текущей страницы: {e}")

# ================================
# Основная функция запуска парсера
# ================================
def run_scraper(config):
    global results, total_vacancies_checked
    results = []
    total_vacancies_checked = 0
    start_time = time.perf_counter()

    options = uc.ChromeOptions()
    options.add_argument(f"--user-data-dir={config['chrome_profile_path']}")
    options.add_argument("--disable-webrtc")
    options.add_argument("--disable-features=WebRtcHideLocalIpsWithMdns")
    options.add_argument("--disable-udp")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--v=0")
    options.add_argument("--disable-blink-features=AutomationControlled")

    if config.get("chrome_binary_location"):
        options.binary_location = config["chrome_binary_location"]

    service = Service(config["chromedriver_path"])
    try:
        driver = uc.Chrome(options=options, service=service)
    except Exception as e:
        logging.error(f"Ошибка запуска браузера: {e}")
        return

    driver.get("https://www.linkedin.com/login")
    logging.info("Ожидаем ручной вход в систему...")
    try:
        WebDriverWait(driver, 60).until(
            lambda d: ("feed" in d.current_url or "linkedin.com/feed" in d.current_url)
        )
        logging.info("Вход выполнен успешно.")
    except Exception as e:
        logging.error("Ошибка при входе в систему. Проверьте логин вручную.")
        driver.quit()
        return

    try:
        # Перебираем все ключевые слова для поиска из настроенного списка (или используем пользовательское)
        search_keywords = SEARCH_KEYWORDS
        if config.get("keyword") and config["keyword"] not in search_keywords:
            search_keywords = [config["keyword"]] + search_keywords
        
        for keyword in search_keywords:
            logging.info(f"=== Начинаем поиск по ключевому слову: {keyword} ===")
            
            # Формируем URL для поиска
            direct_url = (
                "https://www.linkedin.com/jobs/search/"
                f"?keywords={keyword.replace(' ', '%20')}"
                f"&location={config['search_country'].replace(' ', '%20')}"
            )
            driver.get(direct_url)
            time.sleep(3)
            wait = WebDriverWait(driver, 30)

            current_page = 1
            max_pages = 3  # Ограничиваем количество страниц для каждого ключевого слова
            
            while current_page <= max_pages:
                logging.info(f"=== Обработка страницы {current_page} для ключевого слова '{keyword}' ===")
                parse_current_page(driver, wait, start_time, config)
                elapsed_time = round(time.perf_counter() - start_time, 2)
                save_results_to_file_with_calculations(results, config["output_file_path"], elapsed_time)

                next_page_number = current_page + 1
                next_button_xpath = f"//button[@aria-label='Page {next_page_number}']"
                next_buttons = driver.find_elements(By.XPATH, next_button_xpath)
                if not next_buttons:
                    logging.info(f"Страница {next_page_number} не найдена. Пагинация завершена.")
                    break

                next_buttons[0].click()
                current_page += 1
                time.sleep(3)

        elapsed_time = round(time.perf_counter() - start_time, 2)
        save_results_to_file_with_calculations(results, config["output_file_path"], elapsed_time)
        
        # Отправляем итоговый отчет
        if results:
            summary_message = (
                f"📊 <b>Завершен поиск вакансий на LinkedIn</b>\n\n"
                f"Проверено вакансий: {total_vacancies_checked}\n"
                f"Найдено подходящих: {len(results)}\n"
                f"Процент совпадений: {len(results)/total_vacancies_checked*100:.1f}%\n"
                f"Время выполнения: {elapsed_time/60:.1f} минут\n\n"
                f"Ключевые слова поиска: {', '.join(search_keywords)}"
            )
            
            # Создаем финальные графики
            final_bar_chart = create_bar_chart(results)
            final_skills_chart = create_skills_chart(results)
            
            final_images = []
            if final_bar_chart:
                final_images.append(final_bar_chart)
            if final_skills_chart:
                final_images.append(final_skills_chart)
                
            send_telegram_message(config["telegram_bot_token"], config["telegram_chat_id"], summary_message, images=final_images)
        
        logging.info("Пагинация завершена. Финальные результаты сохранены.")
    except Exception as e:
        logging.error(f"Глобальная ошибка при поиске вакансий: {e}")
        elapsed_time = round(time.perf_counter() - start_time, 2)
        save_results_to_file_with_calculations(results, config["output_file_path"], elapsed_time)
    finally:
        driver.quit()
        logging.info("Браузер закрыт.")
        if config.get("shutdown_on_finish"):
            logging.info("Скрипт завершён. Выключение компьютера через 30 секунд.")
            os.system("shutdown /s /t 30")

def start_scraper_thread(config):
    thread = threading.Thread(target=run_scraper, args=(config,))
    thread.start()

# ================================
# Интерфейс (Tkinter)
# ================================
def create_gui():
    root = tk.Tk()
    root.title("LinkedIn Job Scraper")

    search_country_var = tk.StringVar(value=DEFAULT_SEARCH_COUNTRY)
    keyword_var = tk.StringVar(value=DEFAULT_KEYWORD)
    output_file_var = tk.StringVar(value=DEFAULT_OUTPUT_FILE_PATH)
    telegram_bot_token_var = tk.StringVar(value=DEFAULT_TELEGRAM_BOT_TOKEN)
    telegram_chat_id_var = tk.StringVar(value=DEFAULT_TELEGRAM_CHAT_ID)
    chromedriver_path_var = tk.StringVar(value=DEFAULT_CHROMEDRIVER_PATH)
    chrome_profile_path_var = tk.StringVar(value=DEFAULT_CHROME_PROFILE_PATH)
    chrome_binary_location_var = tk.StringVar(value=DEFAULT_CHROME_BINARY_LOCATION)
    shutdown_var = tk.BooleanVar(value=False)

    tk.Label(root, text="Search Country:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=search_country_var, width=40).grid(row=0, column=1, padx=5, pady=5)
    tk.Label(root, text="Primary Keyword:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=keyword_var, width=40).grid(row=1, column=1, padx=5, pady=5)
    
    # Добавляем информацию о предустановленных ключевых словах
    tk.Label(root, text="Примечание: помимо указанного, будет выполнен поиск\nпо дополнительным ключевым словам", 
             justify="left").grid(row=2, column=0, columnspan=2, sticky="w", padx=5, pady=5)
    
    tk.Label(root, text="Output Excel File:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=output_file_var, width=40).grid(row=3, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse", command=lambda: output_file_var.set(
        filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    )).grid(row=3, column=2, padx=5, pady=5)
    tk.Label(root, text="Telegram Bot Token:").grid(row=4, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=telegram_bot_token_var, width=40).grid(row=4, column=1, padx=5, pady=5)
    tk.Label(root, text="Telegram Chat ID:").grid(row=5, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=telegram_chat_id_var, width=40).grid(row=5, column=1, padx=5, pady=5)
    tk.Label(root, text="ChromeDriver Path:").grid(row=6, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=chromedriver_path_var, width=40).grid(row=6, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse", command=lambda: chromedriver_path_var.set(
        filedialog.askopenfilename(filetypes=[("Executable", "*.exe")])
    )).grid(row=6, column=2, padx=5, pady=5)
    tk.Label(root, text="Chrome Profile Path:").grid(row=7, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=chrome_profile_path_var, width=40).grid(row=7, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse", command=lambda: chrome_profile_path_var.set(
        filedialog.askdirectory()
    )).grid(row=7, column=2, padx=5, pady=5)
    tk.Label(root, text="Chrome Binary Location:").grid(row=8, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(root, textvariable=chrome_binary_location_var, width=40).grid(row=8, column=1, padx=5, pady=5)
    tk.Button(root, text="Browse", command=lambda: chrome_binary_location_var.set(
        filedialog.askopenfilename(filetypes=[("Chrome Executable", "*.exe")])
    )).grid(row=8, column=2, padx=5, pady=5)
    shutdown_checkbox = tk.Checkbutton(root, text="Выключить компьютер после завершения работы скрипта", variable=shutdown_var)
    shutdown_checkbox.grid(row=9, column=0, columnspan=3, pady=5)

    def on_start():
        config = {
            "search_country": search_country_var.get(),
            "keyword": keyword_var.get(),
            "output_file_path": output_file_var.get(),
            "telegram_bot_token": telegram_bot_token_var.get(),
            "telegram_chat_id": telegram_chat_id_var.get(),
            "chromedriver_path": chromedriver_path_var.get(),
            "chrome_profile_path": chrome_profile_path_var.get(),
            "chrome_binary_location": chrome_binary_location_var.get(),
            "shutdown_on_finish": shutdown_var.get()
        }
        if not config["output_file_path"]:
            messagebox.showerror("Ошибка", "Укажите путь для сохранения Excel файла.")
            return
        root.destroy()
        start_scraper_thread(config)

    tk.Button(root, text="Start Scraper", command=on_start, bg="green", fg="white").grid(row=10, column=0, columnspan=3, pady=10)
    root.mainloop()

if __name__ == "__main__":
    create_gui()